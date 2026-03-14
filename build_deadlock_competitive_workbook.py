from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
HERO_LIST_URL = "https://www.dexerto.com/wikis/deadlock/heroes/"
OUTPUT_PATH = ROOT / "excel" / "deadlock_competitive_system.xlsx"
ROLE_SOURCE_PATH = ROOT / "excel" / "Character Roles.xlsx"
ROLE_DEF_PATH = ROOT / "excel" / "Deadlock Roles.xlsx"
HEADERS = {"User-Agent": "Mozilla/5.0"}


@dataclass
class HeroPage:
    name: str
    category: str
    weapon: str
    tier: str
    url: str


@dataclass
class Ability:
    hero: str
    name: str
    description: str
    tags: list[str]


def normalize_name(name: str) -> str:
    return (
        name.replace("&amp;", "&")
        .replace("The Doorman", "Doorman")
        .replace("Page", "Paige")
        .strip()
    )


def name_from_slug(slug: str) -> str:
    mapping = {
        "grey-talon": "Grey Talon",
        "lady-geist": "Lady Geist",
        "mcginnis": "McGinnis",
        "mo-krill": "Mo & Krill",
        "the-doorman": "Doorman",
    }
    if slug in mapping:
        return mapping[slug]
    return slug.replace("-", " ").title()


def load_role_map() -> dict[str, str]:
    wb = load_workbook(ROLE_SOURCE_PATH, data_only=True)
    ws = wb.active
    role_map: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        hero, role = row[:2]
        if hero and role:
            role_map[normalize_name(str(hero))] = str(role).strip()
    return role_map


def load_role_definitions() -> list[tuple[str, str]]:
    wb = load_workbook(ROLE_DEF_PATH, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        role, meaning = row[:2]
        if role and meaning:
            rows.append((str(role).strip(), str(meaning).strip()))
    return rows


def get_soup(url: str) -> BeautifulSoup:
    response = requests.get(url, headers=HEADERS, timeout=30)
    response.raise_for_status()
    return BeautifulSoup(response.text, "html.parser")


def scrape_hero_index() -> list[HeroPage]:
    soup = get_soup(HERO_LIST_URL)
    seen: set[str] = set()
    heroes: list[HeroPage] = []

    for link in soup.find_all("a", href=True):
        href = link["href"]
        if not href.startswith("/wikis/deadlock/heroes/"):
            continue
        if href in {"/wikis/deadlock/heroes/", "/wikis/deadlock/heroes/\\"}:
            continue
        if href in seen:
            continue

        text = " ".join(link.get_text(" ", strip=True).split())
        if not text:
            continue

        parts = text.split()
        if len(parts) < 4:
            continue

        slug = href.strip("/").split("/")[-1]
        name = normalize_name(name_from_slug(slug))
        name_tokens = name.replace("&", "").split()
        category_index = len(name_tokens)
        if len(parts) <= category_index + 2:
            continue
        category = parts[category_index]
        weapon = " ".join(parts[category_index + 1 : -1])
        tier = parts[-1]

        seen.add(href)
        clean_href = href.rstrip("\\")
        heroes.append(
            HeroPage(
                name=name,
                category=category,
                weapon=weapon,
                tier=tier,
                url=f"https://www.dexerto.com{clean_href}",
            )
        )

    heroes.sort(key=lambda hero: hero.name)
    return heroes


def scrape_abilities(hero: HeroPage) -> list[Ability]:
    soup = get_soup(hero.url)
    heading = soup.find(
        lambda tag: tag.name == "h2"
        and f"{hero.name} Abilities" in " ".join(tag.get_text(" ", strip=True).split())
    )
    if not heading:
        raise ValueError(f"Could not find ability table for {hero.name}")

    table = heading.find_next("table")
    if not table or not table.tbody:
        raise ValueError(f"Could not parse ability table for {hero.name}")

    abilities: list[Ability] = []
    for row in table.tbody.find_all("tr", recursive=False):
        cells = row.find_all("td", recursive=False)
        if len(cells) < 2:
            continue

        raw_name = " ".join(cells[0].stripped_strings)
        ability_name = re.sub(r"^\(\d+\)\s*", "", raw_name).strip()
        description = " ".join(cells[1].stripped_strings).strip()
        if not ability_name or not description:
            continue

        abilities.append(
            Ability(
                hero=hero.name,
                name=ability_name,
                description=description,
                tags=classify_ability(ability_name, description),
            )
        )

    if len(abilities) != 4:
        raise ValueError(f"Expected 4 abilities for {hero.name}, found {len(abilities)}")
    return abilities


def has_any(text: str, words: Iterable[str]) -> bool:
    return any(word in text for word in words)


def classify_ability(name: str, description: str) -> list[str]:
    text = f"{name} {description}".lower()
    tags: set[str] = set()

    damage_words = {
        "damage",
        "attack",
        "bullet",
        "spirit power",
        "projectile",
        "shoot",
        "shot",
        "beam",
        "burn",
        "explod",
        "strike",
        "slash",
        "stabs",
        "barrage",
        "bomb",
        "fire rate",
        "weapon damage",
    }
    pick_words = {
        "stun",
        "sleep",
        "silence",
        "trap",
        "root",
        "hook",
        "pull",
        "drag",
        "grab",
        "immobil",
        "swap",
        "knock",
        "taunt",
        "slow",
        "restrain",
        "disarm",
        "curse",
    }
    tank_words = {
        "resist",
        "invulner",
        "shield",
        "barrier",
        "armor",
        "bullet resist",
        "spirit resist",
        "damage reduction",
        "block",
        "parry",
        "durable",
        "health",
        "regenerate",
    }
    utility_words = {
        "ally",
        "allies",
        "team",
        "teleport",
        "reveal",
        "vision",
        "buff",
        "debuff",
        "slow",
        "silence",
        "summon",
        "clone",
        "ward",
        "heal",
        "barrier",
        "move speed",
        "cooldown",
        "stamina",
        "dispel",
        "cleanse",
        "invisible",
        "cloak",
    }
    engage_words = {
        "charge",
        "dash",
        "leap",
        "lunge",
        "teleport",
        "launch",
        "jump",
        "burrow",
        "rush",
        "blink",
        "pounce",
    }
    sustain_words = {
        "heal",
        "healing",
        "regenerate",
        "restore health",
        "lifesteal",
        "drain health",
        "steal health",
        "shield",
        "invulner",
        "immune",
        "revive",
    }

    if has_any(text, damage_words):
        tags.add("Damage")
    if has_any(text, pick_words):
        tags.add("Pick")
    if has_any(text, tank_words):
        tags.add("Tank")
    if has_any(text, utility_words):
        tags.add("Utility")
    if has_any(text, engage_words):
        tags.add("Engage")
    if has_any(text, sustain_words):
        tags.add("Sustain")

    if not tags:
        tags.add("Utility")
    return sorted(tags)


def autosize(ws) -> None:
    for col_cells in ws.columns:
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 42)


def role_lookup_formula(hero_ref: str) -> str:
    return f'IFERROR(VLOOKUP({hero_ref},Heroes!$A:$C,3,FALSE),"")'


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(heroes: list[HeroPage], hero_abilities: dict[str, list[Ability]], role_map: dict[str, str]) -> Workbook:
    wb = Workbook()
    hero_end_row = len(heroes) + 1
    ability_end_row = sum(len(v) for v in hero_abilities.values()) + 1

    ws_heroes = wb.active
    ws_heroes.title = "Heroes"
    ws_heroes.append(
        ["Hero", "Category", "Baseline Role", "Weapon", "Tier", "A1", "A2", "A3", "A4"]
    )
    for hero in heroes:
        abilities = hero_abilities[hero.name]
        ws_heroes.append(
            [
                hero.name,
                hero.category,
                role_map.get(hero.name, ""),
                hero.weapon,
                hero.tier,
                abilities[0].name,
                abilities[1].name,
                abilities[2].name,
                abilities[3].name,
            ]
        )
    style_header(ws_heroes)
    ws_heroes.freeze_panes = "A2"
    autosize(ws_heroes)

    ws_tags = wb.create_sheet("Ability Tags")
    ws_tags.append(["Hero", "Ability", "Tag", "Tag Group", "Description"])
    for hero in heroes:
        for ability in hero_abilities[hero.name]:
            for tag in ability.tags:
                tag_group = tag if tag in {"Tank", "Pick", "Damage", "Utility"} else (
                    "Pick" if tag == "Engage" else "Utility"
                )
                ws_tags.append([hero.name, ability.name, tag, tag_group, ability.description])
    style_header(ws_tags)
    ws_tags.freeze_panes = "A2"
    autosize(ws_tags)

    ws_roles = wb.create_sheet("Role Output")
    ws_roles.append(["Hero", "Tank", "Pick", "Damage", "Utility", "Engage", "Sustain", "Role"])
    for row_num, hero in enumerate(heroes, start=2):
        lookup = role_lookup_formula(f"$A{row_num}")
        ability_hero_range = f"'Ability Tags'!$A$2:$A${ability_end_row}"
        ability_group_range = f"'Ability Tags'!$D$2:$D${ability_end_row}"
        ability_tag_range = f"'Ability Tags'!$C$2:$C${ability_end_row}"
        ws_roles.append(
            [
                hero.name,
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_group_range},"Tank")+IF({lookup}="Tank",2,IF({lookup}="Hybrid",1,0))',
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_group_range},"Pick")+IF({lookup}="Initiator/Catch",2,0)',
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_group_range},"Damage")+IF(OR({lookup}="Carry",{lookup}="Off-Carry"),2,IF({lookup}="Hybrid",1,0))',
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_group_range},"Utility")+IF(OR({lookup}="Support",{lookup}="Comfort"),2,IF({lookup}="Hybrid",2,0))',
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_tag_range},"Engage")',
                f'=COUNTIFS({ability_hero_range},$A{row_num},{ability_tag_range},"Sustain")',
                f'=CHOOSE(MATCH(MAX(B{row_num}:E{row_num}),B{row_num}:E{row_num},0),"Tank","Pick","Damage","Utility")',
            ]
        )
    style_header(ws_roles)
    ws_roles.freeze_panes = "A2"
    autosize(ws_roles)

    ws_pools = wb.create_sheet("Hero Pools")
    ws_pools["A1"] = "Baseline Role Pools"
    ws_pools["A1"].font = Font(bold=True, size=14)
    baseline_headers = ["Carry", "Off-Carry", "Tank", "Initiator/Catch", "Support", "Hybrid", "Comfort"]
    for idx, header in enumerate(baseline_headers, start=1):
        col = get_column_letter(idx)
        ws_pools[f"{col}2"] = header
        ws_pools[f"{col}3"] = f'=SORT(FILTER(Heroes!$A$2:$A${hero_end_row},Heroes!$C$2:$C${hero_end_row}={col}$2,""))'
    style_header(ws_pools, row=2)

    ws_pools["J1"] = "Derived Coverage Pools"
    ws_pools["J1"].font = Font(bold=True, size=14)
    derived_headers = ["Tank", "Pick", "Damage", "Utility", "Engage", "Sustain"]
    for offset, header in enumerate(derived_headers, start=10):
        col = get_column_letter(offset)
        ws_pools[f"{col}2"] = header
        source_col = get_column_letter(offset - 8)
        ws_pools[f"{col}3"] = f'=SORT(FILTER(\'Role Output\'!$A$2:$A${hero_end_row},\'Role Output\'!${source_col}$2:${source_col}${hero_end_row}>0,""))'
    style_header(ws_pools, row=2)
    autosize(ws_pools)

    ws_draft = wb.create_sheet("Draft Builder")
    ws_draft["A1"] = "6-Player Draft Builder"
    ws_draft["A1"].font = Font(bold=True, size=14)
    ws_draft.append(
        ["Player", "Hero", "Baseline Role", "Derived Role", "Tank", "Pick", "Damage", "Utility", "Engage", "Sustain"]
    )
    players = [f"Player {i}" for i in range(1, 7)]
    for idx, player in enumerate(players, start=3):
        ws_draft[f"A{idx}"] = player
        ws_draft[f"C{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},Heroes!$A$2:$C${hero_end_row},3,FALSE),""))'
        ws_draft[f"D{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},8,FALSE),""))'
        ws_draft[f"E{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},2,FALSE),0))'
        ws_draft[f"F{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},3,FALSE),0))'
        ws_draft[f"G{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},4,FALSE),0))'
        ws_draft[f"H{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},5,FALSE),0))'
        ws_draft[f"I{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},6,FALSE),0))'
        ws_draft[f"J{idx}"] = f'=IF($B{idx}="","",IFERROR(VLOOKUP($B{idx},\'Role Output\'!$A$2:$H${hero_end_row},7,FALSE),0))'

    style_header(ws_draft, row=2)
    hero_validation = DataValidation(type="list", formula1=f"=Heroes!$A$2:$A${hero_end_row}", allow_blank=True)
    ws_draft.add_data_validation(hero_validation)
    hero_validation.add("B3:B8")

    ws_draft["L2"] = "Lineup Summary"
    ws_draft["L2"].font = Font(bold=True)
    summaries = [
        ("Total Tank", "=SUM(E3:E8)"),
        ("Total Pick", "=SUM(F3:F8)"),
        ("Total Damage", "=SUM(G3:G8)"),
        ("Total Utility", "=SUM(H3:H8)"),
        ("Total Engage", "=SUM(I3:I8)"),
        ("Total Sustain", "=SUM(J3:J8)"),
        ("Core Roles Covered?", '=IF(MIN(M3:M6)>0,"YES","NO")'),
        ("Engage + Sustain + Carry?", '=IF(AND(M7>0,M8>0,M5>0),"YES","NO")'),
        ("Unique Heroes?", '=IF(COUNTA(B3:B8)=COUNTA(UNIQUE(FILTER(B3:B8,B3:B8<>""))),"YES","DUPLICATES")'),
    ]
    for row_num, (label, formula) in enumerate(summaries, start=3):
        ws_draft[f"L{row_num}"] = label
        ws_draft[f"M{row_num}"] = formula

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    ws_draft.conditional_formatting.add("M9:M11", CellIsRule(operator="equal", formula=['"YES"'], fill=green_fill))
    ws_draft.conditional_formatting.add("M9:M11", CellIsRule(operator="notEqual", formula=['"YES"'], fill=red_fill))
    ws_draft.freeze_panes = "A3"
    autosize(ws_draft)

    ws_notes = wb.create_sheet("Notes")
    ws_notes.append(["Source", "Details"])
    ws_notes.append(
        [
            "Dexerto Deadlock Wiki",
            "Current hero list and ability descriptions were scraped from Dexerto's Deadlock hero pages.",
        ]
    )
    ws_notes.append(
        [
            "Steam guide",
            "The referenced Steam guide currently reflects the older 22-hero roster and was used as the original instruction source, but the workbook uses Dexerto so all 38 local roster entries can be filled.",
        ]
    )
    ws_notes.append(
        [
            "Local role files",
            "Baseline roles and role definitions were loaded from excel/Character Roles.xlsx and excel/Deadlock Roles.xlsx.",
        ]
    )
    for role, meaning in load_role_definitions():
        ws_notes.append([role, meaning])
    style_header(ws_notes)
    autosize(ws_notes)

    return wb


def main() -> None:
    role_map = load_role_map()
    heroes = scrape_hero_index()
    hero_abilities = {hero.name: scrape_abilities(hero) for hero in heroes}
    wb = build_workbook(heroes, hero_abilities, role_map)
    wb.save(OUTPUT_PATH)
    print(f"Workbook created: {OUTPUT_PATH}")
    print(f"Heroes: {len(heroes)}")
    print(f"Abilities: {sum(len(v) for v in hero_abilities.values())}")


if __name__ == "__main__":
    main()
